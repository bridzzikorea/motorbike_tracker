import os
import time
import os
import string
import pandas as pd
from pathlib import Path

from datetime import datetime
from openpyxl import Workbook, load_workbook, worksheet

import util.error_log.errors as errors
import util.os.path as path_util


### 테스트가 필요함 ###

def is_col_letter(s: str) -> bool:
    """
    주어진 문자열이 A, B, C 같은 알파벳만(한 문자 이상)으로 이루어져 있는지 간단히 체크하는 함수
    """
    return s.isalpha()

def col_letter_to_number(col_letter): 
    """
    [주의!] google sheet에 접근할 때 사용하는 index이다.
    """
    if not is_col_letter(col_letter):
        raise ValueError(f"입력값은 문자 데이터가 아닙니다: {col_letter}")

    col_letter = col_letter.upper()
    result = 0
    for c in col_letter:
        result = result * 26 + (ord(c) - ord('A')) + 1
    return result

def col_number_to_letter(col_num: int) -> str:
    """
    예: 1 -> A, 2 -> B, ..., 26 -> Z,
       27 -> AA, 28 -> AB, ...
    """
    if not isinstance(col_num, int):
        raise ValueError(f"입력값은 정수형 숫자 데이터가 아닙니다: {col_num}")
    if col_num <= 0:
        raise ValueError(f"숫자는 0보다 큰 수를 입력해야합니다. {col_num}")
        
    letters = []
    while col_num > 0:
        remainder = (col_num - 1) % 26
        letters.append(chr(remainder + ord('A')))
        col_num = (col_num - 1) // 26
    letters = ''.join(reversed(letters))        
    
    if not is_col_letter(letters):
        raise ValueError("컬럼 문자열을 변환 할 수 없습니다.")    
    return letters

def get_min_of_last_data_rows(sheet: worksheet.worksheet.Worksheet, key_col_letters) -> int:
    """
    여러 키 컬럼 중 '가장 아래까지 데이터가 있는 행 번호'를 각각 구한 뒤,
    그 중 가장 작은 값을 반환.

    예) key_col_letters가 ["A", "B"]이고
        - A컬럼은 마지막 데이터가 50행,
        - B컬럼은 마지막 데이터가 30행 이라면,
      두 컬럼 모두 데이터가 있는 행은 최대 30행이므로, 30을 반환.
    """
    if not key_col_letters:
        return 0

    last_data_rows = []
    max_possible = sheet.max_row

    for col_letter in key_col_letters:
        row_found = 0
        for row_idx in range(max_possible, 0, -1):
            cell_value = sheet[f"{col_letter}{row_idx}"].value
            if cell_value not in (None, ""):
                row_found = row_idx
                break
        last_data_rows.append(row_found)

    # 모두 비어있으면 0, 아니면 가장 작은 행 번호
    if all(r == 0 for r in last_data_rows):
        return 0
    return min(r for r in last_data_rows if r > 0)

def get_max_row_in_col_range(sheet: worksheet.worksheet.Worksheet, start_letter: str, end_letter: str) -> int:
    """
    start_letter부터 end_letter까지, 각 컬럼 중
    하나라도 데이터가 있는 가장 아래쪽 행 번호의 최대값을 구한다.

    즉, 열 범위 내 어느 컬럼에라도 마지막으로 데이터가 존재하는 행을 찾고,
    그 중 가장 큰 행 번호를 반환.
    """
    max_found = 0
    sheet_max = sheet.max_row
    for col_code in range(ord(start_letter), ord(end_letter) + 1):
        col_letter = chr(col_code)
        for row_idx in range(sheet_max, 0, -1):
            cell_value = sheet[f"{col_letter}{row_idx}"].value
            if cell_value not in (None, ""):
                if row_idx > max_found:
                    max_found = row_idx
                break
    return max_found

class ExcelSheet:
    """엑셀(.xlsx) 및 CSV 파일의 읽기/쓰기 기능을 제공하는 유틸리티 클래스.
    
    Pandas와 openpyxl을 활용하여 파일을 DataFrame으로 불러오거나 저장하며,
    예외 처리와 대용량 데이터에 대한 최적화 옵션을 포함한다.
    """
    
    def __init__(self, spreadsheet_path: str):
        """
        엑셀 스프레드시트 선택 초기화

        :param spreadsheet_name: 액세스할 스프레드시트 이름
        """        
        file_path = Path(spreadsheet_path)
        if not path_util.is_valid_path(spreadsheet_path):
            file_path.parent.mkdir(parents=True, exist_ok=True)
            file_path.touch()
            print(f"파일이 없어 빈 파일을 생성합니다 : {file_path}")        
        
        self.spreadsheet_path = spreadsheet_path
        self.workbook = load_workbook(self.spreadsheet_path, keep_vba=False)
    
    
    def load_sheet(self, sheet_name: str) -> worksheet.worksheet.Worksheet:
        """
        특정 워크시트를 가져옵니다.
        """
        return self.workbook[sheet_name]

    def load_as_fetched_data(
        self,
        sheet_name: str,
        start_col_letter: str,
        end_col_letter: str,
        key_col_letters=None
    ):
        """
        - 키 컬럼이 있으면: 모든 키 컬럼이 공통으로 값이 있는 마지막 행(= 각 키 컬럼의 last row 중 최소값)까지만 로드.
        - 키 컬럼이 없으면: start_col_letter~end_col_letter 범위 안에 있는 모든 열의 마지막 데이터 행(= 모든 열의 last row 중 최대값)까지 로드.
        """
        if key_col_letters is None:
            key_col_letters = []

        sheet = self.load_sheet(sheet_name)
        if sheet is None:
            raise ValueError("sheet 로드 과정에서 None 데이터가 들어왔습니다")

        # 인자 검증
        if not is_col_letter(start_col_letter):
            raise ValueError(f"start_col_letter에 알파벳이 아닌 문자 데이터가 입력되었습니다: {start_col_letter}")
        if not is_col_letter(end_col_letter):
            raise ValueError(f"end_col_letter에 알파벳이 아닌 문자 데이터가 입력되었습니다: {end_col_letter}")
        for k in key_col_letters:
            if not is_col_letter(k):
                raise ValueError(f"key_col_letter에 알파벳이 아닌 문자 데이터가 입력되었습니다: {k}")

        start_col_letter = start_col_letter.upper()
        end_col_letter = end_col_letter.upper()
        key_col_letters = [letter.upper() for letter in key_col_letters]

        # 키 컬럼이 start~end 범위에 들어가는지 체크
        for k_col in key_col_letters:
            if not (start_col_letter <= k_col <= end_col_letter):
                raise ValueError(f"key_col_letter가 start_col~end_col 범위를 벗어났습니다: {k_col}")

        if len(key_col_letters) == 0:
            # 키 컬럼이 없으면: start~end 열 범위 내에서, 가장 아래쪽(= 최대) 행 번호를 찾는다.
            max_row = get_max_row_in_col_range(sheet, start_col_letter, end_col_letter)
            if max_row == 0:
                print("Fail: 지정된 열 범위 내에 데이터가 전혀 없습니다.")
                raise ValueError("EmptyDataError")
        else:
            # 키 컬럼이 있으면: 각 키 컬럼마다 last row를 찾고, 그 중 최소값까지만 로드.
            min_last_row_in_keys = get_min_of_last_data_rows(sheet, key_col_letters)
            if min_last_row_in_keys == 0:
                print("Fail: 키 컬럼 중 하나라도 전부 비어있어, 불러올 행이 없습니다.")
                raise ValueError("EmptyDataError")
            max_row = min_last_row_in_keys

        # 범위 문자열
        cell_range = f"{start_col_letter}1:{end_col_letter}{max_row}"
        raw_cells = sheet[cell_range]

        sheet_data = []
        for row in raw_cells:
            row_data = []
            for cell in row:
                row_data.append(cell.value if cell.value is not None else "")
            sheet_data.append(row_data)

        # 열 길이 맞추기
        col_len = ord(end_col_letter) - ord(start_col_letter) + 1
        for row_index, row in enumerate(sheet_data):
            while len(row) < col_len:
                sheet_data[row_index].append('')

        print("[ load_as_fetched_data ]")
        if len(sheet_data) == 0:
            print("Fail: 엑셀 시트에서 데이터가 비어 있습니다. 시트와 범위를 확인해주세요.")
            raise ValueError("EmptyDataError")

        if len(sheet_data) == 1:
            for row in sheet_data:
                for val in row:
                    if val == "#N/A":
                        print("Fail: #N/A 값이 존재합니다.")
                        raise ValueError("EmptyDataError")
            print("Success: (헤더만 존재)")
        else:
            print("Success: 엑셀 시트에서 데이터를 성공적으로 가져왔습니다.")

        return sheet_data

    def load_as_dataframe(
        self,
        sheet_name: str,
        start_col_letter: str,
        end_col_letter: str,
        key_cols=None
    ):
        if key_cols is None:
            key_cols = []

        load_data = self.load_as_fetched_data(sheet_name, start_col_letter, end_col_letter, key_cols)

        col_length = len(load_data[0])
        for index, row in enumerate(load_data):
            while len(load_data[index]) < col_length:
                load_data[index].append("")

        print(f"[ load_as_dataframe ] row count: {len(load_data)}")
        if len(load_data) == 1:
            return pd.DataFrame(load_data[1:], columns=load_data[0])
        else:
            if len(load_data[0]) == len(load_data[1]):
                df_sheet = pd.DataFrame(load_data[1:], columns=load_data[0])
            else:
                raise ValueError("Columns and data length do not match.")
            return df_sheet

    def load_one_line(self, sheet_name: str, start_col_letter: str, end_col_letter: str):
        """
        구글 시트에서 사용하던 'load_one_line' 함수를 openpyxl 기반으로 변환한 예시.
        (헤더는 1행, 실제 데이터는 2행 가정)
        """
        sheet = self.load_sheet(sheet_name)
        if sheet is None:
            raise ValueError("Sheet를 불러오지 못했습니다.")

        if not is_col_letter(start_col_letter):
            raise ValueError(f"start_col_letter에 알파벳이 아닌 문자 데이터가 입력되었습니다: {start_col_letter}")
        if not is_col_letter(end_col_letter):
            raise ValueError(f"end_col_letter에 알파벳이 아닌 문자 데이터가 입력되었습니다: {end_col_letter}")

        start_col_letter = start_col_letter.upper()
        end_col_letter = end_col_letter.upper()

        # A1 ~ B2 형태로 2행까지만 읽어옴 (헤더 + 실제 데이터 1줄)
        cell_range = f"{start_col_letter}1:{end_col_letter}2"
        raw_cells = sheet[cell_range]  # ((cellObj, cellObj, ...), (cellObj, cellObj, ...))

        # 2차원 리스트 변환
        sheet_data = []
        for row in raw_cells:
            row_data = []
            for cell in row:
                row_data.append(cell.value if cell.value is not None else "")
            sheet_data.append(row_data)

        # 최소 2행은 존재해야 실제 데이터가 있다고 봄 (헤더 + 실제값)
        if len(sheet_data) < 2:
            print("Fail: 엑셀 시트에서 데이터가 비어 있거나, 2행 데이터가 없습니다.")
            raise ValueError("EmptyDataError")

        print("Success: 엑셀 시트에서 데이터를 가져왔습니다.")

        # 딕셔너리 생성
        oneline_dict = {}
        header_row = sheet_data[0]
        value_row = sheet_data[1]

        # 만약 value_row가 header_row보다 짧으면 빈 문자열로 채워줌
        while len(value_row) < len(header_row):
            value_row.append('')

        for col_index, header in enumerate(header_row):
            value = value_row[col_index]
            oneline_dict[header] = value

        return oneline_dict
    
    def save(self, save_path=None):        
        if save_path == None:
            self.workbook.save(self.spreadsheet_path)
        else:
            self.workbook.save(save_path)
    
    # ----------------------------------------------------------------------
    # 1) vlookup_update
    # ----------------------------------------------------------------------
    def vlookup_update(self, sheet_name, key_value, key_col_letter, start_col_letter, data: list):
        """
        구글 시트에서 사용하던 vlookup_update 함수를
        openpyxl 스타일로 변환한 예시.

        data: 리스트(2차배열 형태)
        key_value: 매칭할 키
        key_col_letter: 매칭 대상 컬럼 (ex: 'A')
        start_col_letter: 매칭된 행에 대해, 이 열부터 data를 쓴다 (ex: 'C')
        """

        if not data:
            raise ValueError("Data list is empty. Update not performed.")
        if not is_col_letter(start_col_letter):
            raise ValueError(f"start_col_letter가 알파벳이 아닙니다: {start_col_letter}")
        if len(data) == 0:
            raise ValueError(f"data 개수가 0입니다. 확인 부탁드립니다.")
        if isinstance(data, list):
            data = [[value] for value in data]

        sheet = self.load_sheet(sheet_name)
        if sheet is None:
            raise ValueError("시트를 찾지 못했습니다.")

        # 1) key_col_letter 로드 → 해당 열을 pandas로 전체 로드
        key_data = self.load_as_dataframe(sheet_name, key_col_letter, key_col_letter, key_col_letter)
        # key_data: 한 컬럼만 있는 DF (ex: row0=header, row1~ = 값)
        # key_value 와 매칭되는 row 인덱스를 찾는다

        start_row_number = 0
        for index, key in enumerate(key_data[key_data.columns[0]], start=1):
            # enumerate(...) start=1 => DF의 첫 데이터가 1행이 됨
            # but 주의: 실제 엑셀의 1행은 헤더인지 여부는 load_as_dataframe 구현에 따라 다름
            if key == key_value:
                start_row_number = index + 1  # +1은 header를 제외한 실제 데이터행
                break
        
        if start_row_number == 0:
            raise ValueError(f"키 {key_value}를 {key_col_letter}에서 찾지 못했습니다.")

        # data 크기 파악
        data_row_count = len(data)         # 행 수
        data_column_count = len(data[0])   # 열 수

        # 2) 쓰기 시작할 열/행 계산
        start_col_number = col_letter_to_number(start_col_letter)
        end_col_number = start_col_number + data_column_count - 1
        end_col_letter = col_number_to_letter(end_col_number)

        # ex) A10~C10 (만약 data_row_count=1이면 10행 한 줄만)
        end_row_number = start_row_number + data_row_count - 1

        # 3) 실제 엑셀에 쓰기
        # openpyxl에는 sheet.range(...)가 없으므로, 반복문으로 cell.value 할당
        for i in range(data_row_count):
            for j in range(data_column_count):
                cell_row = start_row_number + j
                cell_col = start_col_number + i
                sheet.cell(row=cell_row, column=cell_col).value = data[i][j]

        # 4) 저장
        self.workbook.save(self.spreadsheet_path)

    # ----------------------------------------------------------------------
    # 2) update (여러 행을 아래쪽에 추가)
    # ----------------------------------------------------------------------
    def update(self, sheet_name, output_rows, start_col_letter):
        """
        - 구글 시트 코드에서는 start_col_letter 열의 끝(빈칸)부터 이어서 쓰는 로직.
        - openpyxl에서는 '마지막으로 값이 들어있는 행'을 찾은 뒤 그 다음 행부터 기록.
        """

        if not output_rows:
            raise ValueError(f"{sheet_name}에 출력할 데이터가 없습니다.")
        if not is_col_letter(start_col_letter):
            raise ValueError(f"start_col_letter가 알파벳이 아닙니다: {start_col_letter}")

        sheet = self.load_sheet(sheet_name)
        if sheet is None:
            raise ValueError("시트를 찾지 못했습니다.")

        start_col_number = col_letter_to_number(start_col_letter)
        data_row_count = len(output_rows)
        data_column_count = len(output_rows[0])

        # 1) 'start_col_letter' 열에서 마지막으로 값이 들어있는 행 찾기
        #    구글 시트의 col_values().length와 비슷한 로직
        last_nonempty_row = 0
        for row_idx in range(1, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=start_col_number).value
            if val is not None and val != "":
                last_nonempty_row = row_idx
        
        # 마지막 값이 들어있는 행 아래로 쓰면 되므로
        start_row_number = last_nonempty_row + 1

        end_row_number = start_row_number + data_row_count - 1
        end_col_number = start_col_number + data_column_count - 1

        # 2) 필요한 경우, sheet.insert_rows()로 행을 확장 (구글 시트의 add_rows와 유사)
        #    만약 end_row_number > sheet.max_row, 모자란 만큼 insert
        if end_row_number > sheet.max_row:
            extra_rows = end_row_number - sheet.max_row
            # openpyxl은 특정 위치에 행을 삽입: sheet.insert_rows(idx, amount)
            # 여기서는 sheet.max_row+1 위치(마지막)부터 extra_rows만큼 삽입
            sheet.insert_rows(sheet.max_row + 1, amount=extra_rows)

        # 3) 실제 쓰기
        for i, row_data in enumerate(output_rows):
            for j, value in enumerate(row_data):
                cell_row = start_row_number + i
                cell_col = start_col_number + j
                sheet.cell(row=cell_row, column=cell_col).value = value

        # 4) 저장
        self.workbook.save(self.spreadsheet_path)

    # ----------------------------------------------------------------------
    # 3) update_oneline (1줄만 추가)
    # ----------------------------------------------------------------------
    def update_oneline(self, sheet_name, oneline_data, start_col_letter):
        """
        oneline_data: 1차원 리스트(한 행)
        - 구글 시트 원본은 'start_col_letter 열'의 마지막 빈칸 찾아서 그 행에 씀
        """

        if not oneline_data:
            raise ValueError("oneline_data is empty. Update not performed.")
        if not is_col_letter(start_col_letter):
            raise ValueError(f"start_col_letter가 알파벳이 아닙니다: {start_col_letter}")

        sheet = self.load_sheet(sheet_name)
        if sheet is None:
            raise ValueError("시트를 찾지 못했습니다.")

        start_col_number = col_letter_to_number(start_col_letter)
        data_row_count = 1  # 한 줄
        data_column_count = len(oneline_data)

        # 1) 마지막으로 값이 들어있는 행 찾기
        last_nonempty_row = 0
        for row_idx in range(1, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=start_col_number).value
            if val is not None and val != "":
                last_nonempty_row = row_idx
        
        start_row_number = last_nonempty_row + 1
        end_row_number = start_row_number
        end_col_number = start_col_number + data_column_count - 1

        # 2) 필요한 경우 insert_rows
        if end_row_number > sheet.max_row:
            extra_rows = end_row_number - sheet.max_row
            sheet.insert_rows(sheet.max_row + 1, amount=extra_rows)

        # 3) 실제 쓰기
        for j, value in enumerate(oneline_data):
            cell_col = start_col_number + j
            sheet.cell(row=start_row_number, column=cell_col).value = value

        # 4) 저장
        self.workbook.save(self.spreadsheet_path)
